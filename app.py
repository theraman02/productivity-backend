from fastapi import FastAPI, Depends, HTTPException, Form, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import SessionLocal
from models import Employee, WeeklyScore, User, Organization, OrganizationMembership
from scoring import calculate_productivity
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import secrets
import hashlib
import uuid

app = FastAPI(title="Employee Productivity Tracker")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://productivity-tracker-three.vercel.app",
        "http://localhost:3000"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

sessions = {}


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def create_session(user_id: int, organization_id: str, role: str) -> str:
    token = secrets.token_urlsafe(32)
    sessions[token] = {
        'user_id': user_id,
        'organization_id': organization_id,
        'role': role,
        'expires': datetime.now() + timedelta(days=7)
    }
    return token


def verify_session(token: str):
    if token not in sessions:
        return None
    session = sessions[token]
    if datetime.now() > session['expires']:
        del sessions[token]
        return None
    return session


def get_current_user(authorization: str = Header(None), db: Session = Depends(get_db)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Not authenticated")

    token = authorization.replace("Bearer ", "")
    session = verify_session(token)

    if not session:
        raise HTTPException(
            status_code=401, detail="Invalid or expired session")

    return session


# ============= AUTHENTICATION =============

@app.post("/auth/register")
def register(
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    existing_user = db.query(User).filter(
        (User.username == username) | (User.email == email)
    ).first()

    if existing_user:
        raise HTTPException(
            status_code=400, detail="Username or email already exists")

    # Create user
    organization_id = str(uuid.uuid4())
    hashed_pw = hash_password(password)

    user = User(
        username=username,
        email=email,
        password_hash=hashed_pw,
        primary_organization_id=organization_id
    )
    db.add(user)
    db.flush()

    # Create organization
    org = Organization(
        id=organization_id,
        name=f"{username}'s Organization",
        owner_id=user.id
    )
    db.add(org)

    # Create membership
    membership = OrganizationMembership(
        user_id=user.id,
        organization_id=organization_id,
        role="admin"
    )
    db.add(membership)

    db.commit()
    db.refresh(user)

    token = create_session(user.id, organization_id, "admin")

    return {
        "message": "User registered successfully",
        "token": token,
        "user": {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": "admin",
            "organization_id": organization_id
        }
    }


@app.post("/auth/login")
def login(
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.username == username).first()

    if not user or user.password_hash != hash_password(password):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Get user's primary organization and role
    membership = db.query(OrganizationMembership).filter(
        OrganizationMembership.user_id == user.id,
        OrganizationMembership.organization_id == user.primary_organization_id
    ).first()

    # FIX: Handle case where membership doesn't exist
    if not membership:
        # Check if user has any memberships at all
        any_membership = db.query(OrganizationMembership).filter(
            OrganizationMembership.user_id == user.id
        ).first()

        if any_membership:
            # Use the first available membership
            membership = any_membership
            # Update user's primary organization
            user.primary_organization_id = any_membership.organization_id
            db.commit()
        else:
            raise HTTPException(
                status_code=500,
                detail="User has no organization memberships. Please contact support."
            )

    role = membership.role
    token = create_session(user.id, membership.organization_id, role)

    return {
        "message": "Login successful",
        "token": token,
        "user": {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": role,
            "organization_id": membership.organization_id
        }
    }


@app.post("/auth/logout")
def logout(token: str = Form(...)):
    if token in sessions:
        del sessions[token]
    return {"message": "Logged out successfully"}


@app.get("/auth/me")
def get_current_user_info(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == current_user['user_id']).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Get all organizations user belongs to
    memberships = db.query(OrganizationMembership, Organization).join(
        Organization, OrganizationMembership.organization_id == Organization.id
    ).filter(OrganizationMembership.user_id == user.id).all()

    organizations = [
        {
            "id": org.id,
            "name": org.name,
            "role": membership.role,
            "is_primary": org.id == user.primary_organization_id
        }
        for membership, org in memberships
    ]

    return {
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "role": current_user['role'],
        "organization_id": current_user['organization_id'],
        "organizations": organizations
    }


@app.delete("/auth/account")
def delete_account(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == current_user['user_id']).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Delete user's primary organization (if they're the owner)
    org = db.query(Organization).filter(
        Organization.id == user.primary_organization_id,
        Organization.owner_id == user.id
    ).first()

    if org:
        # Delete all employees in this organization
        db.query(Employee).filter(Employee.organization_id == org.id).delete()
        # Delete all scores in this organization
        db.query(WeeklyScore).filter(
            WeeklyScore.organization_id == org.id).delete()
        # Delete organization
        db.delete(org)

    # Delete all memberships
    db.query(OrganizationMembership).filter(
        OrganizationMembership.user_id == user.id).delete()

    # Delete user
    db.delete(user)
    db.commit()

    # Delete session
    token = None
    for t, s in sessions.items():
        if s['user_id'] == user.id:
            token = t
            break
    if token:
        del sessions[token]

    return {"message": "Account deleted successfully"}


# ============= ORGANIZATION SWITCHING =============

@app.post("/organizations/switch")
def switch_organization(
    organization_id: str = Form(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Check if user is member of this organization
    membership = db.query(OrganizationMembership).filter(
        OrganizationMembership.user_id == current_user['user_id'],
        OrganizationMembership.organization_id == organization_id
    ).first()

    if not membership:
        raise HTTPException(
            status_code=403, detail="You are not a member of this organization")

    # Update session
    for token, session in sessions.items():
        if session['user_id'] == current_user['user_id']:
            session['organization_id'] = organization_id
            session['role'] = membership.role

            return {
                "message": "Organization switched successfully",
                "organization_id": organization_id,
                "role": membership.role
            }

    raise HTTPException(status_code=500, detail="Session not found")


# ============= EMPLOYEE ENDPOINTS =============

@app.get("/")
def root():
    return {"status": "Backend running successfully"}


@app.get("/employees")
def get_employees(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(Employee).filter(
        Employee.organization_id == current_user['organization_id']
    ).all()


@app.get("/employees/{employee_id}")
def get_employee(employee_id: int, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(
        Employee.id == employee_id,
        Employee.organization_id == current_user['organization_id']
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    return employee


@app.post("/employees")
def create_employee(
    name: str,
    department: str,
    role: str,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can create employees")

    employee = Employee(
        name=name,
        department=department,
        role=role,
        organization_id=current_user['organization_id']
    )
    db.add(employee)
    db.commit()
    db.refresh(employee)
    return employee


@app.put("/employees/{employee_id}")
def update_employee(
    employee_id: int,
    name: str = None,
    department: str = None,
    role: str = None,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can update employees")

    employee = db.query(Employee).filter(
        Employee.id == employee_id,
        Employee.organization_id == current_user['organization_id']
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    if name:
        employee.name = name
    if department:
        employee.department = department
    if role:
        employee.role = role

    db.commit()
    db.refresh(employee)
    return employee


@app.delete("/employees/{employee_id}")
def delete_employee(
    employee_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can delete employees")

    employee = db.query(Employee).filter(
        Employee.id == employee_id,
        Employee.organization_id == current_user['organization_id']
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    db.query(WeeklyScore).filter(
        WeeklyScore.employee_id == employee_id).delete()
    db.delete(employee)
    db.commit()
    return {"message": "Employee deleted successfully"}


# ============= SCORE ENDPOINTS =============

@app.post("/scores")
def add_weekly_score(
    employee_id: int,
    week: str,
    task_completion: float,
    speed: float,
    professionalism: float,
    activity: float,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can add scores")

    employee = db.query(Employee).filter(
        Employee.id == employee_id,
        Employee.organization_id == current_user['organization_id']
    ).first()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    productivity = calculate_productivity(
        task_completion, speed, professionalism, activity
    )

    score = WeeklyScore(
        employee_id=employee_id,
        week=week,
        task_completion=task_completion,
        speed=speed,
        professionalism=professionalism,
        activity=activity,
        productivity_score=productivity,
        organization_id=current_user['organization_id']
    )

    db.add(score)
    db.commit()
    db.refresh(score)
    return score


@app.get("/scores")
def get_scores(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(WeeklyScore).filter(
        WeeklyScore.organization_id == current_user['organization_id']
    ).all()


@app.delete("/scores/{score_id}")
def delete_score(
    score_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can delete scores")

    score = db.query(WeeklyScore).filter(
        WeeklyScore.id == score_id,
        WeeklyScore.organization_id == current_user['organization_id']
    ).first()

    if not score:
        raise HTTPException(status_code=404, detail="Score not found")

    db.delete(score)
    db.commit()
    return {"message": "Score deleted successfully"}

# ============= TEAM MANAGEMENT ENDPOINTS =============


@app.get("/team/members")
def get_team_members(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can view team members")

    # Get all memberships for current organization
    memberships = db.query(OrganizationMembership, User).join(
        User, OrganizationMembership.user_id == User.id
    ).filter(OrganizationMembership.organization_id == current_user['organization_id']).all()

    return [
        {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": membership.role,
            "created_at": user.created_at
        }
        for membership, user in memberships
    ]


@app.post("/team/invite")
def invite_user_to_team(
    username_or_email: str = Form(...),
    role: str = Form("viewer"),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can invite users")

    # Find user by username or email
    user = db.query(User).filter(
        (User.username == username_or_email) | (
            User.email == username_or_email)
    ).first()

    if not user:
        raise HTTPException(
            status_code=404, detail="User not found. They need to register first.")

    # Check if already a member
    existing = db.query(OrganizationMembership).filter(
        OrganizationMembership.user_id == user.id,
        OrganizationMembership.organization_id == current_user['organization_id']
    ).first()

    if existing:
        raise HTTPException(
            status_code=400, detail="User is already a member of this organization")

    # Validate role
    if role not in ['admin', 'viewer']:
        raise HTTPException(status_code=400, detail="Invalid role")

    # Create membership
    membership = OrganizationMembership(
        user_id=user.id,
        organization_id=current_user['organization_id'],
        role=role
    )

    db.add(membership)
    db.commit()

    return {
        "message": f"User {user.username} invited successfully",
        "user": {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": role
        }
    }


@app.post("/team/members")
def create_team_member(
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form("viewer"),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can add team members")

    # Check if username/email exists
    existing = db.query(User).filter(
        (User.username == username) | (User.email == email)
    ).first()

    if existing:
        # User exists, invite them instead
        existing_membership = db.query(OrganizationMembership).filter(
            OrganizationMembership.user_id == existing.id,
            OrganizationMembership.organization_id == current_user['organization_id']
        ).first()

        if existing_membership:
            raise HTTPException(
                status_code=400, detail="User is already a member of this organization")

        # Add them to the organization
        membership = OrganizationMembership(
            user_id=existing.id,
            organization_id=current_user['organization_id'],
            role=role
        )
        db.add(membership)
        db.commit()

        return {
            "id": existing.id,
            "username": existing.username,
            "email": existing.email,
            "role": role,
            "message": "Existing user added to organization"
        }

    # Validate role
    if role not in ['admin', 'viewer']:
        raise HTTPException(status_code=400, detail="Invalid role")

    # Create new user without their own organization
    hashed_pw = hash_password(password)
    new_user = User(
        username=username,
        email=email,
        password_hash=hashed_pw,
        primary_organization_id=current_user['organization_id']
    )

    db.add(new_user)
    db.flush()

    # Add to current organization
    membership = OrganizationMembership(
        user_id=new_user.id,
        organization_id=current_user['organization_id'],
        role=role
    )
    db.add(membership)

    db.commit()
    db.refresh(new_user)

    return {
        "id": new_user.id,
        "username": new_user.username,
        "email": new_user.email,
        "role": role
    }


@app.delete("/team/members/{user_id}")
def remove_team_member(
    user_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user['role'] != 'admin':
        raise HTTPException(
            status_code=403, detail="Only admins can remove team members")

    if user_id == current_user['user_id']:
        raise HTTPException(status_code=400, detail="Cannot remove yourself")

    # Remove membership
    membership = db.query(OrganizationMembership).filter(
        OrganizationMembership.user_id == user_id,
        OrganizationMembership.organization_id == current_user['organization_id']
    ).first()

    if not membership:
        raise HTTPException(
            status_code=404, detail="User is not a member of this organization")

    db.delete(membership)
    db.commit()

    return {"message": "Team member removed successfully"}


# ============= EXPORT ENDPOINTS =============

@app.get("/export/excel/{week}")
def export_to_excel(week: str, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    scores = db.query(WeeklyScore).filter(
        WeeklyScore.week == week,
        WeeklyScore.organization_id == current_user['organization_id']
    ).all()

    if not scores:
        raise HTTPException(
            status_code=404, detail=f"No scores found for week {week}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Week {week}"

    ws['A1'] = f'Weekly Productivity Report - {week}'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')

    headers = ['Employee ID', 'Employee Name', 'Task', 'Speed',
               'Professional', 'Activity', 'Productivity Score']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4",
                                end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    for row, score in enumerate(scores, start=4):
        employee = db.query(Employee).filter(
            Employee.id == score.employee_id).first()
        ws.cell(row=row, column=1).value = score.employee_id
        ws.cell(row=row, column=2).value = employee.name if employee else 'Unknown'
        ws.cell(row=row, column=3).value = score.task_completion
        ws.cell(row=row, column=4).value = score.speed
        ws.cell(row=row, column=5).value = score.professionalism
        ws.cell(row=row, column=6).value = score.activity
        ws.cell(row=row, column=7).value = score.productivity_score
        ws.cell(row=row, column=7).font = Font(bold=True)

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=weekly_report_{week}.xlsx"}
    )


@app.get("/export/pdf/{week}")
def export_to_pdf(week: str, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    scores = db.query(WeeklyScore).filter(
        WeeklyScore.week == week,
        WeeklyScore.organization_id == current_user['organization_id']
    ).all()

    if not scores:
        raise HTTPException(
            status_code=404, detail=f"No scores found for week {week}")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph(
        f"<b>Weekly Productivity Report - {week}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))

    data = [['Employee', 'Task', 'Speed', 'Prof.', 'Activity', 'Score']]

    for score in scores:
        employee = db.query(Employee).filter(
            Employee.id == score.employee_id).first()
        data.append([
            employee.name if employee else 'Unknown',
            str(score.task_completion),
            str(score.speed),
            str(score.professionalism),
            str(score.activity),
            str(score.productivity_score)
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=weekly_report_{week}.pdf"}
    )


@app.post("/email/report")
def email_report(
    week: str = Form(...),
    recipient_email: str = Form(...),
    smtp_server: str = Form("smtp.gmail.com"),
    smtp_port: int = Form(587),
    sender_email: str = Form(...),
    sender_password: str = Form(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        scores = db.query(WeeklyScore).filter(
            WeeklyScore.week == week,
            WeeklyScore.organization_id == current_user['organization_id']
        ).all()

        if not scores:
            raise HTTPException(
                status_code=404, detail=f"No scores for week {week}")

        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = f'Weekly Productivity Report - {week}'

        body = f"""
        <html>
        <body>
        <h2>Weekly Productivity Report - {week}</h2>
        <p>Please find the weekly productivity report attached.</p>
        <br>
        <table border="1" cellpadding="5">
        <tr style="background-color: #4472C4; color: white;">
            <th>Employee</th>
            <th>Task</th>
            <th>Speed</th>
            <th>Professional</th>
            <th>Activity</th>
            <th>Score</th>
        </tr>
        """

        for score in scores:
            employee = db.query(Employee).filter(
                Employee.id == score.employee_id).first()
            body += f"""
            <tr>
                <td>{employee.name if employee else 'Unknown'}</td>
                <td>{score.task_completion}</td>
                <td>{score.speed}</td>
                <td>{score.professionalism}</td>
                <td>{score.activity}</td>
                <td><b>{score.productivity_score}</b></td>
            </tr>
            """

        body += """
        </table>
        <br>
        <p>Best regards,<br>Productivity Tracker System</p>
        </body>
        </html>
        """

        msg.attach(MIMEText(body, 'html'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()

        return {"message": f"Report emailed successfully to {recipient_email}"}

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error sending email: {str(e)}")
